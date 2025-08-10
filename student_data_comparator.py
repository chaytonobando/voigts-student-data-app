#!/usr/bin/env python3
"""
🎓 Student Data Comparison Tool
Compares exported data from AI PDF extractor with another Excel sheet,
matching student names and highlighting the matching rows.

Features:
- Load and compare two Excel files
- Fuzzy name matching for better accuracy
- Highlight matching rows in output
- Generate comparison reports
- Export results to Excel with highlighting
"""

import os
import sys
import pandas as pd
import logging
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime
import re
from difflib import SequenceMatcher

# Import required libraries
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from fuzzywuzzy import fuzz, process
except ImportError as e:
    print(f"❌ Missing required library: {e}")
    print("Please install required packages:")
    print("pip install pandas openpyxl fuzzywuzzy python-levenshtein")
    sys.exit(1)


class StudentDataComparator:
    """Main class for comparing student data between Excel files"""
    
    def __init__(self, log_level=logging.INFO):
        """Initialize the comparator with logging"""
        self.setup_logging(log_level)
        self.ai_extractor_data = None
        self.comparison_data = None
        self.matches = []
        self.unmatched_ai = []
        self.unmatched_comparison = []
        
    def setup_logging(self, log_level):
        """Setup logging configuration"""
        log_filename = f"student_comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
        log_path = Path("logs") / log_filename
        log_path.parent.mkdir(exist_ok=True)
        
        logging.basicConfig(
            level=log_level,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_path),
                logging.StreamHandler(sys.stdout)
            ]
        )
        self.logger = logging.getLogger(__name__)
        self.logger.info("🎓 Student Data Comparator initialized")
    
    def load_ai_extractor_data(self, file_path: str, sheet_name: str = None) -> pd.DataFrame:
        """Load data from AI PDF extractor output with automatic sheet detection"""
        self.logger.info(f"📊 Loading AI extractor data from: {file_path}")
        
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"AI extractor file not found: {file_path}")
            
            # Get all available sheets first
            excel_file = pd.ExcelFile(file_path)
            available_sheets = excel_file.sheet_names
            self.logger.info(f"📄 Available sheets: {available_sheets}")
            
            sheet_to_use = None
            
            # If sheet name is specified, try to use it
            if sheet_name:
                if sheet_name in available_sheets:
                    sheet_to_use = sheet_name
                    self.logger.info(f"✅ Using specified sheet: {sheet_name}")
                else:
                    self.logger.warning(f"⚠️ Specified sheet '{sheet_name}' not found. Available sheets: {available_sheets}")
            
            # If no sheet specified or specified sheet not found, auto-detect
            if not sheet_to_use:
                # Try default "Extracted Data" first
                if "Extracted Data" in available_sheets:
                    sheet_to_use = "Extracted Data"
                    self.logger.info(f"✅ Using default 'Extracted Data' sheet")
                else:
                    # Priority keywords for AI extracted data
                    keywords = ['data', 'student', 'extract', 'form', 'opt', 'sheet1']
                    
                    # Try to find sheet with relevant keywords
                    for keyword in keywords:
                        for sheet in available_sheets:
                            if keyword.lower() in sheet.lower():
                                sheet_to_use = sheet
                                self.logger.info(f"✅ Auto-detected sheet with keyword '{keyword}': {sheet}")
                                break
                        if sheet_to_use:
                            break
                    
                    # If no keyword match, use first sheet
                    if not sheet_to_use:
                        sheet_to_use = available_sheets[0]
                        self.logger.info(f"🔄 Using first available sheet: {sheet_to_use}")
            
            # Load the selected sheet
            self.ai_extractor_data = pd.read_excel(file_path, sheet_name=sheet_to_use)
            
            self.logger.info(f"✅ Loaded {len(self.ai_extractor_data)} rows from AI extractor (sheet: {sheet_to_use})")
            self.logger.info(f"📋 Columns: {list(self.ai_extractor_data.columns)}")
            
            return self.ai_extractor_data
            
        except Exception as e:
            self.logger.error(f"❌ Error loading AI extractor data: {e}")
            raise
    
    def load_comparison_data(self, file_path: str, sheet_name: str = None) -> pd.DataFrame:
        """Load comparison Excel file with automatic sheet detection"""
        self.logger.info(f"📊 Loading comparison data from: {file_path}")
        
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"Comparison file not found: {file_path}")
            
            # Get all available sheets first
            excel_file = pd.ExcelFile(file_path)
            available_sheets = excel_file.sheet_names
            self.logger.info(f"📄 Available sheets: {available_sheets}")
            
            sheet_to_use = None
            
            # If sheet name is specified, try to use it
            if sheet_name:
                if sheet_name in available_sheets:
                    sheet_to_use = sheet_name
                    self.logger.info(f"✅ Using specified sheet: {sheet_name}")
                else:
                    self.logger.warning(f"⚠️ Specified sheet '{sheet_name}' not found. Available sheets: {available_sheets}")
            
            # If no sheet specified or specified sheet not found, auto-detect
            if not sheet_to_use:
                # Priority keywords for comparison data (broader than AI data since templates vary more)
                keywords = ['template', 'data', 'student', 'form', 'comparison', 'traversa', 'opt', 'sheet1']
                
                # Try to find sheet with relevant keywords
                for keyword in keywords:
                    for sheet in available_sheets:
                        if keyword.lower() in sheet.lower():
                            sheet_to_use = sheet
                            self.logger.info(f"✅ Auto-detected sheet with keyword '{keyword}': {sheet}")
                            break
                    if sheet_to_use:
                        break
                
                # If no keyword match, use first sheet
                if not sheet_to_use:
                    sheet_to_use = available_sheets[0]
                    self.logger.info(f"🔄 Using first available sheet: {sheet_to_use}")
            
            # Load the selected sheet
            self.comparison_data = pd.read_excel(file_path, sheet_name=sheet_to_use)
            
            self.logger.info(f"✅ Loaded {len(self.comparison_data)} rows from comparison file (sheet: {sheet_to_use})")
            self.logger.info(f"📋 Columns: {list(self.comparison_data.columns)}")
            
            return self.comparison_data
            
        except Exception as e:
            self.logger.error(f"❌ Error loading comparison data: {e}")
            raise
    
    def detect_name_columns(self, df: pd.DataFrame) -> List[str]:
        """Automatically detect columns that likely contain student names"""
        name_indicators = [
            'name', 'student', 'full_name', 'firstname', 'lastname', 
            'first_name', 'last_name', 'student_name', 'pupil', 'learner'
        ]
        
        potential_columns = []
        first_name_col = None
        last_name_col = None
        
        # First pass: look for specific name indicators
        for col in df.columns:
            col_lower = str(col).lower().replace(' ', '_')
            
            # Skip obviously wrong columns
            if any(skip in col_lower for skip in ['unnamed', 'index', 'id', 'number_of']):
                continue
            
            # Check for first name
            if any(indicator in col_lower for indicator in ['first_name', 'firstname', 'first']):
                first_name_col = col
                potential_columns.append(col)
                continue
            
            # Check for last name
            if any(indicator in col_lower for indicator in ['last_name', 'lastname', 'last']):
                last_name_col = col
                potential_columns.append(col)
                continue
            
            # Check for other name indicators
            for indicator in name_indicators:
                if indicator in col_lower and col not in potential_columns:
                    # Verify this column contains actual name-like data
                    sample_values = df[col].dropna().head(5).astype(str)
                    # Check if values look like names - handle common name formats including "Last, First"
                    if any(self._is_name_like(val) for val in sample_values):
                        potential_columns.append(col)
                        break
        
        # If we have both first and last name columns, prioritize the combination
        if first_name_col and last_name_col:
            self.logger.info(f"🔍 Found separate first/last name columns: {first_name_col}, {last_name_col}")
            # Remove individual columns and add them back in the right order
            if first_name_col in potential_columns:
                potential_columns.remove(first_name_col)
            if last_name_col in potential_columns:
                potential_columns.remove(last_name_col)
            # Add them back at the beginning
            potential_columns.insert(0, last_name_col)
            potential_columns.insert(0, first_name_col)
        
        # If no obvious name columns found, look for columns with string data that looks like names
        if not potential_columns:
            for col in df.columns:
                if df[col].dtype == 'object':  # String/object columns
                    sample_values = df[col].dropna().head(10).astype(str)
                    # Check if values look like names using improved detection
                    name_like_count = sum(1 for val in sample_values if self._is_name_like(val))
                    if name_like_count >= len(sample_values) * 0.7:  # 70% of samples look like names
                        potential_columns.append(col)
        
        self.logger.info(f"🔍 Detected potential name columns: {potential_columns}")
        return potential_columns
    
    def _is_name_like(self, val: str) -> bool:
        """Check if a value looks like a name, handling various formats"""
        if not val or len(val.strip()) < 2 or len(val.strip()) > 100:
            return False
        
        val_str = str(val).strip()
        
        # Remove common separators and check if what remains is mostly alphabetic
        cleaned = val_str.replace(' ', '').replace(',', '').replace('.', '').replace('-', '').replace("'", '').replace('*', '')
        
        # Must contain mostly letters
        if not cleaned.isalpha():
            return False
        
        # Common patterns for names
        # "Last, First" format
        if ',' in val_str:
            parts = val_str.split(',')
            if len(parts) == 2:
                last, first = parts
                return (last.strip().replace(' ', '').isalpha() and 
                       first.strip().replace(' ', '').isalpha() and
                       len(last.strip()) > 0 and len(first.strip()) > 0)
        
        # "First Last" or "First Middle Last" format
        words = val_str.split()
        if len(words) >= 1:
            # All words should be mostly alphabetic
            for word in words:
                word_clean = word.replace('.', '').replace("'", '').replace('-', '')
                if not word_clean.isalpha() or len(word_clean) < 1:
                    return False
            return True
        
        return False
    
    def normalize_name(self, name: str) -> str:
        """Normalize a name for better matching"""
        if pd.isna(name) or not name:
            return ""
        
        # Convert to string and normalize
        name_str = str(name).strip()
        
        # Remove extra whitespace and normalize case
        name_str = re.sub(r'\s+', ' ', name_str)
        name_str = name_str.title()  # Proper case
        
        # Remove common prefixes/suffixes that might interfere
        prefixes = ['mr.', 'mrs.', 'ms.', 'dr.', 'prof.']
        suffixes = ['jr.', 'sr.', 'ii', 'iii', 'iv']
        
        words = name_str.lower().split()
        cleaned_words = []
        
        for word in words:
            word_clean = word.rstrip('.,')
            if word_clean not in prefixes and word_clean not in suffixes:
                cleaned_words.append(word_clean.title())
        
        return ' '.join(cleaned_words)
    
    def _create_combined_name(self, row: pd.Series, name_columns: List[str]) -> str:
        """Create a combined name from row data, handling first/last name columns intelligently"""
        combined_parts = []
        
        for col in name_columns:
            if pd.notna(row[col]) and str(row[col]).strip():
                col_value = str(row[col]).strip()
                
                # Skip obviously non-name values
                if col_value.lower() in ['nan', 'none', ''] or col_value.replace('.', '').isdigit():
                    continue
                
                # Clean the value
                cleaned_value = self.normalize_name(col_value)
                if cleaned_value and len(cleaned_value) > 1:
                    combined_parts.append(cleaned_value)
        
        # Join all parts with spaces
        combined_name = ' '.join(combined_parts)
        
        # Final cleanup - remove extra spaces
        combined_name = re.sub(r'\s+', ' ', combined_name).strip()
        
        return combined_name
    
    def fuzzy_match_names(self, name1: str, name2: str, threshold: int = 80) -> Tuple[bool, int]:
        """Check if two names match using fuzzy matching"""
        if not name1 or not name2:
            return False, 0
        
        # Normalize both names
        norm_name1 = self.normalize_name(name1)
        norm_name2 = self.normalize_name(name2)
        
        if not norm_name1 or not norm_name2:
            return False, 0
        
        # Try different matching strategies
        scores = []
        
        # Direct ratio
        scores.append(fuzz.ratio(norm_name1, norm_name2))
        
        # Token sort ratio (handles different word orders)
        scores.append(fuzz.token_sort_ratio(norm_name1, norm_name2))
        
        # Token set ratio (handles partial matches)
        scores.append(fuzz.token_set_ratio(norm_name1, norm_name2))
        
        # Partial ratio (for when one name is contained in another)
        scores.append(fuzz.partial_ratio(norm_name1, norm_name2))
        
        # Use the best score
        best_score = max(scores)
        
        return best_score >= threshold, best_score
    
    def compare_data(self, ai_name_columns: List[str] = None, comparison_name_columns: List[str] = None, 
                    fuzzy_threshold: int = 80) -> Dict[str, Any]:
        """Compare the two datasets and find matching students"""
        self.logger.info("🔍 Starting student data comparison...")
        
        if self.ai_extractor_data is None or self.comparison_data is None:
            raise ValueError("Both datasets must be loaded before comparison")
        
        # Auto-detect name columns if not provided
        if ai_name_columns is None:
            detected_ai_columns = self.detect_name_columns(self.ai_extractor_data)
            # Prioritize student name columns over parent name columns
            ai_name_columns = []
            
            # First, look for exact student name column
            for col in detected_ai_columns:
                col_lower = str(col).lower()
                if 'student name' in col_lower and '*' not in col_lower:
                    ai_name_columns = [col]  # Use only this column
                    break
            
            # If no student name column, prioritize First Name/Last Name over Parent names
            if not ai_name_columns:
                student_first_last = []
                for col in detected_ai_columns:
                    col_lower = str(col).lower()
                    # Prefer columns without "parent" in the name
                    if 'parent' not in col_lower and ('first name' in col_lower or 'last name' in col_lower):
                        student_first_last.append(col)
                
                if student_first_last:
                    ai_name_columns = student_first_last
                elif detected_ai_columns:
                    ai_name_columns = [detected_ai_columns[0]]  # Fallback to first detected
        
        if comparison_name_columns is None:
            comparison_name_columns = self.detect_name_columns(self.comparison_data)
        
        if not ai_name_columns or not comparison_name_columns:
            self.logger.warning("⚠️ Could not detect name columns automatically")
            return {"error": "No name columns detected"}
        
        self.logger.info(f"📝 Using AI extractor name columns: {ai_name_columns}")
        self.logger.info(f"📝 Using comparison name columns: {comparison_name_columns}")
        
        # Reset match results
        self.matches = []
        self.unmatched_ai = []
        self.unmatched_comparison = []
        
        # Create combined name fields for better matching
        ai_names = []
        for idx, row in self.ai_extractor_data.iterrows():
            combined_name = self._create_combined_name(row, ai_name_columns)
            if combined_name.strip():  # Only add if we have a valid name
                ai_names.append({
                    'index': idx,
                    'name': combined_name.strip(),
                    'data': row
                })
        
        comparison_names = []
        for idx, row in self.comparison_data.iterrows():
            combined_name = self._create_combined_name(row, comparison_name_columns)
            if combined_name.strip():  # Only add if we have a valid name
                comparison_names.append({
                    'index': idx,
                    'name': combined_name.strip(),
                    'data': row
                })
        
        self.logger.info(f"📊 Loaded {len(ai_names)} valid AI student names")
        self.logger.info(f"📊 Loaded {len(comparison_names)} valid comparison student names")
        
        # Show some examples of the names we're working with
        if ai_names:
            sample_ai = [student['name'] for student in ai_names[:3]]
            self.logger.info(f"📝 AI name samples: {sample_ai}")
        
        if comparison_names:
            sample_comp = [student['name'] for student in comparison_names[:3]]
            self.logger.info(f"📝 Comparison name samples: {sample_comp}")
        
        # Track which comparison rows have been matched
        matched_comparison_indices = set()
        
        # Find matches
        for ai_student in ai_names:
            best_match = None
            best_score = 0
            
            for comp_student in comparison_names:
                if comp_student['index'] in matched_comparison_indices:
                    continue  # Already matched
                
                is_match, score = self.fuzzy_match_names(ai_student['name'], comp_student['name'], fuzzy_threshold)
                
                if is_match and score > best_score:
                    best_match = comp_student
                    best_score = score
            
            if best_match:
                self.matches.append({
                    'ai_index': ai_student['index'],
                    'comparison_index': best_match['index'],
                    'ai_name': ai_student['name'],
                    'comparison_name': best_match['name'],
                    'match_score': best_score,
                    'ai_data': ai_student['data'],
                    'comparison_data': best_match['data']
                })
                matched_comparison_indices.add(best_match['index'])
            else:
                self.unmatched_ai.append({
                    'index': ai_student['index'],
                    'name': ai_student['name'],
                    'data': ai_student['data']
                })
        
        # Find unmatched comparison students
        for comp_student in comparison_names:
            if comp_student['index'] not in matched_comparison_indices:
                self.unmatched_comparison.append({
                    'index': comp_student['index'],
                    'name': comp_student['name'],
                    'data': comp_student['data']
                })
        
        results = {
            'total_ai_students': len(ai_names),
            'total_comparison_students': len(comparison_names),
            'matches_found': len(self.matches),
            'unmatched_ai': len(self.unmatched_ai),
            'unmatched_comparison': len(self.unmatched_comparison),
            'match_rate': len(self.matches) / len(ai_names) * 100 if ai_names else 0
        }
        
        self.logger.info(f"✅ Comparison complete:")
        self.logger.info(f"   📊 Total AI students: {results['total_ai_students']}")
        self.logger.info(f"   📊 Total comparison students: {results['total_comparison_students']}")
        self.logger.info(f"   ✅ Matches found: {results['matches_found']}")
        self.logger.info(f"   ❌ Unmatched AI: {results['unmatched_ai']}")
        self.logger.info(f"   ❌ Unmatched comparison: {results['unmatched_comparison']}")
        self.logger.info(f"   📈 Match rate: {results['match_rate']:.1f}%")
        
        return results
    
    def export_results(self, output_path: str):
        """Export comparison results to Excel with highlighting"""
        self.logger.info(f"📊 Exporting comparison results to: {output_path}")
        
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                
                # Create summary sheet
                summary_data = [{
                    'Total AI Students': len(self.ai_extractor_data) if self.ai_extractor_data is not None else 0,
                    'Total Comparison Students': len(self.comparison_data) if self.comparison_data is not None else 0,
                    'Matches Found': len(self.matches),
                    'Unmatched AI Students': len(self.unmatched_ai),
                    'Unmatched Comparison Students': len(self.unmatched_comparison),
                    'Match Rate (%)': f"{len(self.matches) / max(1, len(self.ai_extractor_data)) * 100:.1f}" if self.ai_extractor_data is not None else "0.0"
                }]
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name="Summary", index=False)
                
                # Create matches sheet
                if self.matches:
                    matches_data = []
                    for match in self.matches:
                        match_row = {
                            'AI_Name': match['ai_name'],
                            'Comparison_Name': match['comparison_name'],
                            'Match_Score': match['match_score'],
                            'AI_Source_File': match['ai_data'].get('Source File', ''),
                        }
                        
                        # Add AI data columns
                        for col, val in match['ai_data'].items():
                            if col != 'Source File':
                                match_row[f'AI_{col}'] = val
                        
                        # Add comparison data columns
                        for col, val in match['comparison_data'].items():
                            match_row[f'Comp_{col}'] = val
                        
                        matches_data.append(match_row)
                    
                    matches_df = pd.DataFrame(matches_data)
                    matches_df.to_excel(writer, sheet_name="Matches", index=False)
                
                # Create unmatched AI sheet
                if self.unmatched_ai:
                    unmatched_ai_data = [student['data'].to_dict() for student in self.unmatched_ai]
                    unmatched_ai_df = pd.DataFrame(unmatched_ai_data)
                    unmatched_ai_df.to_excel(writer, sheet_name="Unmatched_AI", index=False)
                
                # Create unmatched comparison sheet
                if self.unmatched_comparison:
                    unmatched_comp_data = [student['data'].to_dict() for student in self.unmatched_comparison]
                    unmatched_comp_df = pd.DataFrame(unmatched_comp_data)
                    unmatched_comp_df.to_excel(writer, sheet_name="Unmatched_Comparison", index=False)
            
            # Apply formatting
            self._apply_excel_formatting(output_path)
            
            self.logger.info(f"✅ Results exported successfully to: {output_path}")
            
        except Exception as e:
            self.logger.error(f"❌ Error exporting results: {e}")
            raise
    
    def _apply_excel_formatting(self, file_path: str):
        """Apply formatting and highlighting to the Excel file"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            match_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
            unmatch_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
            
            # Format each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Format headers
                for cell in sheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Apply highlighting based on sheet type
                if sheet_name == "Matches":
                    # Highlight matched rows in green
                    for row in sheet.iter_rows(min_row=2):
                        for cell in row:
                            cell.fill = match_fill
                
                elif "Unmatched" in sheet_name:
                    # Highlight unmatched rows in light red
                    for row in sheet.iter_rows(min_row=2):
                        for cell in row:
                            cell.fill = unmatch_fill
                
                # Auto-adjust column widths
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    sheet.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(file_path)
            self.logger.info("✅ Excel formatting applied successfully")
            
        except Exception as e:
            self.logger.error(f"❌ Error applying Excel formatting: {e}")


def main():
    """Main function to demonstrate usage"""
    print("🎓 Student Data Comparison Tool")
    print("=" * 50)
    
    # Initialize comparator
    comparator = StudentDataComparator()
    
    # Example usage - you'll need to update these paths
    ai_extractor_file = "/Users/chaytonobando/Library/Mobile Documents/com~apple~CloudDocs/Python/AI_PDF_Data_Extractor/ROCORI_380_Files_Extraction.xlsx"
    comparison_file = input("📁 Enter path to comparison Excel file: ").strip()
    
    if not comparison_file:
        print("❌ No comparison file provided")
        return
    
    try:
        # Load data
        comparator.load_ai_extractor_data(ai_extractor_file)
        comparator.load_comparison_data(comparison_file)
        
        # Perform comparison
        results = comparator.compare_data(fuzzy_threshold=80)
        
        if 'error' in results:
            print(f"❌ Error: {results['error']}")
            return
        
        # Export results
        output_file = f"student_comparison_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        comparator.export_results(output_file)
        
        print(f"\n✅ Comparison complete! Results saved to: {output_file}")
        
    except Exception as e:
        print(f"❌ Error: {e}")


if __name__ == "__main__":
    main()
