#!/usr/bin/env python3
"""
🚌 Traversa Data Processor
Processes student data comparison results specifically for Traversa routing software upload.

Features:
- Maintains original import template format for Traversa compatibility
- Removes unmatched students from the dataset
- Updates matched students with AI-extracted information
- Highlights changes between old and new data
- Preserves column structure and formatting
- Generates Traversa-ready Excel file
"""

import os
import sys
import pandas as pd
import logging
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime
import re

# Import required libraries
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    try:
        from openpyxl.comments import Comment
        COMMENTS_AVAILABLE = True
    except ImportError:
        COMMENTS_AVAILABLE = False
        Comment = None
except ImportError as e:
    print(f"❌ Missing required library: {e}")
    print("Please install required packages:")
    print("pip install pandas openpyxl")
    sys.exit(1)

# Import our existing comparator
try:
    from student_data_comparator import StudentDataComparator
    COMPARATOR_AVAILABLE = True
except ImportError as e:
    print(f"❌ Warning: StudentDataComparator not available: {e}")
    COMPARATOR_AVAILABLE = False
    StudentDataComparator = None


class TraversaDataProcessor:
    """Process student data for Traversa routing software upload"""
    
    def __init__(self, log_level=logging.INFO):
        """Initialize the Traversa processor"""
        self.setup_logging(log_level)
        
        if not COMPARATOR_AVAILABLE:
            raise ImportError("StudentDataComparator is not available - cannot initialize TraversaDataProcessor")
        
        self.comparator = StudentDataComparator(log_level)
        self.traversa_data = None
        self.field_mappings = {}
        self.changes_detected = []
        
    def setup_logging(self, log_level):
        """Set up logging configuration"""
        logging.basicConfig(
            level=log_level,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.StreamHandler(sys.stdout),
                logging.FileHandler(f'traversa_processor_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log')
            ]
        )
        self.logger = logging.getLogger(__name__)
        
    def set_field_mappings(self, mappings: Dict[str, str]):
        """
        Set field mappings between AI extracted data and Traversa template columns
        
        Args:
            mappings: Dict mapping AI field names to Traversa column names
                     e.g., {'Student Name': 'Student_Name', 'Grade': 'Grade_Level'}
        """
        self.field_mappings = mappings
        self.logger.info(f"📋 Field mappings set: {mappings}")
        
    def process_for_traversa(self, 
                           ai_extractor_file: str,
                           traversa_template_file: str,
                           output_file: str,
                           fuzzy_threshold: int = 80,
                           auto_map_fields: bool = True) -> Dict[str, Any]:
        """
        Process student data for Traversa upload
        
        Args:
            ai_extractor_file: Path to AI extracted data Excel file
            traversa_template_file: Path to Traversa template/comparison Excel file
            output_file: Path for output Traversa-ready Excel file
            fuzzy_threshold: Fuzzy matching threshold (0-100)
            auto_map_fields: Whether to auto-map similar field names
            
        Returns:
            Dict with processing results and statistics
        """
        self.logger.info("🚌 Starting Traversa data processing...")
        
        try:
            # Step 1: Load and compare data using existing comparator
            self.logger.info("📊 Loading and comparing student data...")
            
            self.comparator.load_ai_extractor_data(ai_extractor_file)
            self.comparator.load_comparison_data(traversa_template_file)
            
            comparison_results = self.comparator.compare_data(fuzzy_threshold=fuzzy_threshold)
            
            if 'error' in comparison_results:
                raise ValueError(f"Comparison failed: {comparison_results['error']}")
            
            # Step 2: Auto-map fields if requested
            if auto_map_fields and not self.field_mappings:
                self._auto_map_fields()
            
            # Step 3: Create Traversa-ready dataset
            traversa_ready_data = self._create_traversa_dataset()
            
            # Step 4: Export to Excel with highlighting
            self._export_traversa_excel(traversa_ready_data, output_file)
            
            # Step 5: Generate processing summary
            results = {
                'total_original_students': len(self.comparator.comparison_data),
                'matched_students': len(self.comparator.matches),
                'removed_students': len(self.comparator.unmatched_comparison),
                'updated_fields': len(self.changes_detected),
                'traversa_ready_count': len(traversa_ready_data),
                'output_file': output_file,
                'field_mappings_used': self.field_mappings,
                'changes_summary': self._get_changes_summary()
            }
            
            self.logger.info("✅ Traversa data processing completed successfully!")
            self.logger.info(f"📊 Results Summary:")
            self.logger.info(f"   🎓 Original students in template: {results['total_original_students']}")
            self.logger.info(f"   ✅ Students matched and kept: {results['matched_students']}")
            self.logger.info(f"   ❌ Students removed (unmatched): {results['removed_students']}")
            self.logger.info(f"   🔄 Field updates detected: {results['updated_fields']}")
            self.logger.info(f"   📋 Final Traversa-ready count: {results['traversa_ready_count']}")
            
            return results
            
        except Exception as e:
            self.logger.error(f"❌ Error processing data for Traversa: {e}")
            raise
    
    def _auto_map_fields(self):
        """Automatically map similar field names between AI data and Traversa template"""
        self.logger.info("🔍 Auto-mapping fields between AI data and Traversa template...")
        
        ai_columns = list(self.comparator.ai_extractor_data.columns)
        traversa_columns = list(self.comparator.comparison_data.columns)
        
        mappings = {}
        
        # Common field mappings (excluding name fields which are handled separately)
        common_mappings = {
            'grade': ['grade', 'grade_level', 'class', 'year'],
            'address': ['address', 'home_address', 'street_address', 'residence'],
            'parent_name': ['parent', 'guardian', 'parent_name', 'guardian_name', 'contact_name'],
            'phone': ['phone', 'telephone', 'contact_number', 'phone_number', 'mobile'],
            'email': ['email', 'email_address', 'contact_email', 'parent_email'],
            'school': ['school', 'school_name', 'institution'],
            'bus_route': ['route', 'bus_route', 'bus_number', 'transport_route'],
            'pickup_time': ['pickup', 'pickup_time', 'collection_time'],
            'dropoff_time': ['dropoff', 'drop_off', 'dropoff_time', 'delivery_time']
        }
        
        # Try to match columns
        for ai_col in ai_columns:
            ai_col_lower = ai_col.lower().replace(' ', '_').replace('-', '_')
            
            # Skip obviously non-useful columns and NAME-related columns
            skip_words = ['unnamed', 'index', 'number_of_students', 'student_name', 'name', 'first_name', 'last_name', 'full_name']
            if any(skip_word in ai_col_lower for skip_word in skip_words):
                continue
            
            best_match = None
            best_score = 0
            
            for traversa_col in traversa_columns:
                traversa_col_lower = traversa_col.lower().replace(' ', '_').replace('-', '_')
                
                # Skip obviously non-useful columns and NAME-related columns
                skip_words_traversa = ['unnamed', 'index', 'id', 'first_name', 'last_name', 'middle_name', 'student_name']
                if any(skip_word in traversa_col_lower for skip_word in skip_words_traversa):
                    continue
                
                # Direct match
                if ai_col_lower == traversa_col_lower:
                    mappings[ai_col] = traversa_col
                    break
                
                # Check common mappings
                for field_type, variations in common_mappings.items():
                    if any(var in ai_col_lower for var in variations) and any(var in traversa_col_lower for var in variations):
                        mappings[ai_col] = traversa_col
                        break
        
        self.field_mappings = mappings
        self.logger.info(f"📋 Auto-mapped {len(mappings)} fields (excluding name fields): {mappings}")
    
    def _normalize_address_for_comparison(self, address: str) -> str:
        """More aggressive address normalization for comparison purposes"""
        if not address:
            return ""
        
        addr = address.lower().strip()
        
        # Remove common suffixes that don't affect address identity
        suffixes_to_remove = [
            ', united states', ' united states', ' usa', ' us',
            ', mn', ' mn', ', minnesota', ' minnesota'
        ]
        
        for suffix in suffixes_to_remove:
            if addr.endswith(suffix):
                addr = addr[:-len(suffix)].strip()
        
        # Normalize directional abbreviations
        directional_replacements = {
            ' north ': ' n ', ' south ': ' s ', ' east ': ' e ', ' west ': ' w ',
            ' n ': ' n ', ' s ': ' s ', ' e ': ' e ', ' w ': ' w ',  # Keep as single letters
        }
        
        for full, abbrev in directional_replacements.items():
            addr = addr.replace(full, abbrev)
        
        # Normalize street types
        street_replacements = {
            ' street': ' st', ' avenue': ' ave', ' road': ' rd', ' drive': ' dr',
            ' boulevard': ' blvd', ' lane': ' ln', ' court': ' ct', ' circle': ' cir',
            ' place': ' pl'
        }
        
        for full, abbrev in street_replacements.items():
            addr = addr.replace(full, abbrev)
        
        # Remove extra punctuation and spaces
        addr = addr.replace(',', ' ').replace('.', '')
        addr = ' '.join(addr.split())  # Normalize spaces
        
        return addr
    
    def _addresses_are_equivalent(self, addr1: str, addr2: str) -> bool:
        """Check if two addresses are equivalent after aggressive normalization"""
        norm1 = self._normalize_address_for_comparison(addr1)
        norm2 = self._normalize_address_for_comparison(addr2)
        
        # Direct match
        if norm1 == norm2:
            return True
        
        # Check if one address is contained in the other (handles missing city/state)
        if norm1 and norm2:
            # Extract house number and street from each
            parts1 = norm1.split()
            parts2 = norm2.split()
            
            if len(parts1) >= 2 and len(parts2) >= 2:
                # Compare house number and first few street components
                house_street1 = ' '.join(parts1[:3])  # House + street + type
                house_street2 = ' '.join(parts2[:3])
                
                # If the core address components match, consider equivalent
                if house_street1 == house_street2:
                    return True
                
                # Also check if one is contained in the other
                if house_street1 in norm2 or house_street2 in norm1:
                    return True
        
        return False

    def _clean_address_value(self, address_value: str) -> str:
        """Clean address value by removing AI export prefixes and normalizing format"""
        if pd.isna(address_value) or not address_value:
            return ""
        
        address_str = str(address_value).strip()
        
        # Remove common AI export prefixes (including variations)
        prefixes_to_remove = [
            ':selected:', ':Selected:', ':SELECTED:',
            'Selected:', 'selected:', 'SELECTED:',
            'Choice:', 'choice:', 'CHOICE:',
            'Option:', 'option:', 'OPTION:',
            'Address:', 'address:', 'ADDRESS:',
            ':choice:', ':Choice:', ':CHOICE:',
            ':option:', ':Option:', ':OPTION:',
            ':address:', ':Address:', ':ADDRESS:'
        ]
        
        for prefix in prefixes_to_remove:
            if address_str.startswith(prefix):
                address_str = address_str[len(prefix):].strip()
                break
        
        # Basic address normalization
        # Remove extra spaces
        address_str = ' '.join(address_str.split())
        
        # Standardize common abbreviations for better comparison
        replacements = {
            ' N ': ' North ',
            ' S ': ' South ', 
            ' E ': ' East ',
            ' W ': ' West ',
            ' St ': ' Street ',
            ' Ave ': ' Avenue ',
            ' Rd ': ' Road ',
            ' Dr ': ' Drive ',
            ' Blvd ': ' Boulevard ',
            ' Ln ': ' Lane ',
            ' Ct ': ' Court ',
            ' Cir ': ' Circle ',
            ' Pl ': ' Place '
        }
        
        for abbrev, full in replacements.items():
            # Handle different cases
            address_str = address_str.replace(abbrev, full)
            address_str = address_str.replace(abbrev.lower(), full.lower())
            address_str = address_str.replace(abbrev.upper(), full.upper())
        
        return address_str.strip()

    def _analyze_transportation_needs(self, transport_text: str) -> str:
        """Analyze transportation text and categorize the needs"""
        if pd.isna(transport_text) or not transport_text:
            return "No Transportation"
        
        transport_str = str(transport_text).lower()
        
        # Check for explicit "no transportation" indicators
        no_transport_indicators = [
            "not need transportation", "decline service", "no transportation",
            "will not need", "do not need", "don't need", "no transport"
        ]
        
        if any(indicator in transport_str for indicator in no_transport_indicators):
            return "No Transportation"
        
        # Check for AM and PM indicators
        has_am = any(indicator in transport_str for indicator in [
            "am route", "home to school", "morning", "to school", "am program"
        ])
        
        has_pm = any(indicator in transport_str for indicator in [
            "pm route", "school to home", "afternoon", "from school", "pm program"
        ])
        
        # Determine transportation category
        if has_am and has_pm:
            return "Both AM & PM"
        elif has_am:
            return "AM Only"
        elif has_pm:
            return "PM Only"
        else:
            # If transportation text exists but doesn't match patterns, assume both
            return "Other/Unclear"

    def _create_traversa_dataset(self) -> pd.DataFrame:
        """Create the Traversa-ready dataset with only matched students and updated information"""
        self.logger.info("🔄 Creating Traversa-ready dataset...")
        
        # Start with the original Traversa template structure
        traversa_data = self.comparator.comparison_data.copy()
        
        # Track which rows to keep (only matched students)
        keep_indices = []
        self.changes_detected = []
        
        for match in self.comparator.matches:
            comparison_index = match['comparison_index']
            keep_indices.append(comparison_index)
            
            # Update the row with AI data
            original_row = traversa_data.loc[comparison_index].copy()
            
            # Special handling for name changes - check if AI name differs from Traversa combined name
            ai_name = match['ai_name']  # This is the name from AI data
            traversa_first = str(original_row.get('First Name', '')).strip()
            traversa_last = str(original_row.get('Last Name', '')).strip()
            traversa_combined = f"{traversa_first} {traversa_last}".strip()
            
            # Check if names are different (case-insensitive comparison)
            if ai_name.lower() != traversa_combined.lower() and ai_name.strip():
                # Names are different - we need to update and track this as a change
                # Split AI name into first/last parts for Traversa format
                ai_name_parts = ai_name.strip().split()
                if len(ai_name_parts) >= 2:
                    new_first = ai_name_parts[0]
                    new_last = ' '.join(ai_name_parts[1:])  # Everything after first word is last name
                elif len(ai_name_parts) == 1:
                    new_first = ai_name_parts[0]
                    new_last = traversa_last  # Keep existing last name if AI only has first name
                else:
                    new_first = traversa_first
                    new_last = traversa_last
                
                # Update First Name if different
                if new_first != traversa_first:
                    self.changes_detected.append({
                        'student': traversa_combined,
                        'field': 'First Name',
                        'category': 'student_name',
                        'old_value': traversa_first,
                        'new_value': new_first,
                        'row_index': comparison_index
                    })
                    traversa_data.loc[comparison_index, 'First Name'] = new_first
                
                # Update Last Name if different  
                if new_last != traversa_last:
                    self.changes_detected.append({
                        'student': traversa_combined,
                        'field': 'Last Name', 
                        'category': 'student_name',
                        'old_value': traversa_last,
                        'new_value': new_last,
                        'row_index': comparison_index
                    })
                    traversa_data.loc[comparison_index, 'Last Name'] = new_last
            
            # Apply other field mappings (excluding name fields)
            for ai_field, traversa_field in self.field_mappings.items():
                if traversa_field in traversa_data.columns and ai_field in match['ai_data']:
                    old_value = traversa_data.loc[comparison_index, traversa_field]
                    new_value = match['ai_data'][ai_field]
                    
                    # Determine field category for color coding and special handling
                    field_lower = traversa_field.lower()
                    is_address_field = any(keyword in field_lower for keyword in ['address', 'street', 'home', 'residence', 'location'])
                    
                    # Special handling for address fields
                    if is_address_field:
                        # Clean both values and check for equivalence
                        clean_old = self._clean_address_value(old_value)
                        clean_new = self._clean_address_value(new_value)
                        
                        # Use enhanced address comparison
                        addresses_equivalent = self._addresses_are_equivalent(clean_old, clean_new)
                        
                        # Only consider it a change if addresses are genuinely different
                        if pd.notna(new_value) and not addresses_equivalent and clean_new:
                            self.changes_detected.append({
                                'student': match['comparison_name'],
                                'field': traversa_field,
                                'category': 'address',
                                'old_value': old_value,
                                'new_value': clean_new,  # Use cleaned value for update
                                'row_index': comparison_index
                            })
                            
                            # Update with cleaned value
                            traversa_data.loc[comparison_index, traversa_field] = clean_new
                    else:
                        # Regular field handling (non-address)
                        if pd.notna(new_value) and str(old_value) != str(new_value):
                            if any(keyword in field_lower for keyword in ['daycare', 'childcare', 'preschool', 'nursery', 'care']):
                                category = 'daycare'
                            else:
                                category = 'general'
                            
                            self.changes_detected.append({
                                'student': match['comparison_name'],
                                'field': traversa_field,
                                'category': category,
                                'old_value': old_value,
                                'new_value': new_value,
                                'row_index': comparison_index
                            })
                            
                            # Update the value
                            traversa_data.loc[comparison_index, traversa_field] = new_value
            
            # Add specific AI columns that should be preserved
            ai_columns_to_preserve = [
                'Please select one of the following *',
                'When do you need transportation (check all that apply)'
            ]
            
            for ai_col in ai_columns_to_preserve:
                if ai_col in match['ai_data'] and pd.notna(match['ai_data'][ai_col]):
                    # Add the column to traversa_data if it doesn't exist
                    if ai_col not in traversa_data.columns:
                        traversa_data[ai_col] = ''
                    
                    # Set the value for this student
                    traversa_data.loc[comparison_index, ai_col] = match['ai_data'][ai_col]
        
        # Filter to keep only matched students
        traversa_ready = traversa_data.loc[keep_indices].copy()
        
        # Reset index for clean output
        traversa_ready = traversa_ready.reset_index(drop=True)
        
        # Add transportation analysis if the column exists
        if 'When do you need transportation (check all that apply)' in traversa_ready.columns:
            self.logger.info("🚌 Analyzing transportation needs for sorting and highlighting...")
            
            # Add transportation category column
            traversa_ready['Transportation_Category'] = traversa_ready['When do you need transportation (check all that apply)'].apply(
                self._analyze_transportation_needs
            )
            
            # Define sorting order for transportation categories
            transport_order = {
                'AM Only': 1,
                'PM Only': 2, 
                'Both AM & PM': 3,
                'Other/Unclear': 4,
                'No Transportation': 5
            }
            
            # Add sort key for ordering
            traversa_ready['_sort_key'] = traversa_ready['Transportation_Category'].map(transport_order)
            
            # Sort by transportation needs
            traversa_ready = traversa_ready.sort_values('_sort_key').reset_index(drop=True)
            
            # Remove the temporary sort key column
            traversa_ready = traversa_ready.drop('_sort_key', axis=1)
            
            self.logger.info(f"✅ Students sorted by transportation needs:")
            transport_counts = traversa_ready['Transportation_Category'].value_counts()
            for category, count in transport_counts.items():
                self.logger.info(f"   🚌 {category}: {count} students")
        
        self.traversa_data = traversa_ready
        
        self.logger.info(f"✅ Created Traversa dataset with {len(traversa_ready)} students")
        self.logger.info(f"🔄 Detected {len(self.changes_detected)} field changes")
        
        return traversa_ready
    
    def _export_traversa_excel(self, data: pd.DataFrame, output_path: str):
        """Export Traversa-ready data to Excel with change highlighting"""
        self.logger.info(f"📊 Exporting Traversa-ready data to: {output_path}")
        
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                
                # Main Traversa data sheet
                data.to_excel(writer, sheet_name="Traversa_Ready_Data", index=False)
                
                # Changes summary sheet
                if self.changes_detected:
                    changes_df = pd.DataFrame(self.changes_detected)
                    changes_df.to_excel(writer, sheet_name="Changes_Summary", index=False)
                
                # Processing summary sheet
                summary_data = [{
                    'Processing_Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'Original_Student_Count': len(self.comparator.comparison_data) if self.comparator.comparison_data is not None else 0,
                    'Matched_Students': len(self.comparator.matches),
                    'Removed_Students': len(self.comparator.unmatched_comparison),
                    'Final_Student_Count': len(data),
                    'Fields_Updated': len(self.changes_detected),
                    'Ready_For_Traversa': 'YES' if len(data) > 0 else 'NO'
                }]
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name="Processing_Summary", index=False)
                
                # Color Legend sheet
                legend_data = [
                    {'Category': 'Address', 'Color': 'Light Blue', 'Description': 'Home address, street address, residence information'},
                    {'Category': 'Student Name', 'Color': 'Light Purple', 'Description': 'Student name, first name, last name, full name'},
                    {'Category': 'Daycare', 'Color': 'Light Green', 'Description': 'Daycare address, childcare, preschool information'},
                    {'Category': 'General', 'Color': 'Light Orange', 'Description': 'All other fields (grade, phone, email, etc.)'},
                    {'Category': 'TRANSPORTATION CATEGORIES', 'Color': '', 'Description': ''},
                    {'Category': 'AM Only', 'Color': 'Light Orange', 'Description': '🌅 Morning transportation only (home to school)'},
                    {'Category': 'PM Only', 'Color': 'Light Indigo', 'Description': '🌆 Afternoon transportation only (school to home)'},
                    {'Category': 'Both AM & PM', 'Color': 'Light Cyan', 'Description': '🔄 Full transportation service (both directions)'},
                    {'Category': 'No Transportation', 'Color': 'Light Red', 'Description': '🚫 No transportation service needed/declined'},
                    {'Category': 'Other/Unclear', 'Color': 'Light Purple', 'Description': '❓ Transportation needs require review'}
                ]
                legend_df = pd.DataFrame(legend_data)
                legend_df.to_excel(writer, sheet_name="Color_Legend", index=False)
            
            # Apply formatting
            self._apply_traversa_formatting(output_path)
            
            self.logger.info("✅ Traversa Excel export completed")
            
        except Exception as e:
            self.logger.error(f"❌ Error exporting Traversa Excel: {e}")
            raise
    
    def _apply_traversa_formatting(self, file_path: str):
        """Apply formatting to highlight changes in the Traversa Excel file with category-specific colors"""
        self.logger.info("🎨 Applying formatting to Traversa Excel file...")
        
        try:
            workbook = openpyxl.load_workbook(file_path)
            
            # Define styles
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            # Category-specific colors for changes
            address_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # Light blue for address
            address_font = Font(color="1976D2", bold=True)
            
            student_name_fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")  # Light purple for student name
            student_name_font = Font(color="7B1FA2", bold=True)
            
            daycare_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")  # Light green for daycare
            daycare_font = Font(color="388E3C", bold=True)
            
            general_fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")  # Light orange for other fields
            general_font = Font(color="CC6600", bold=True)
            
            # Transportation-specific colors
            am_only_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")  # Light orange for AM only
            am_only_font = Font(color="F57C00", bold=True)
            
            pm_only_fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")  # Light indigo for PM only  
            pm_only_font = Font(color="3F51B5", bold=True)
            
            both_transport_fill = PatternFill(start_color="E1F5FE", end_color="E1F5FE", fill_type="solid")  # Light cyan for both
            both_transport_font = Font(color="0277BD", bold=True)
            
            no_transport_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")  # Light red for no transport
            no_transport_font = Font(color="C62828", bold=True)
            
            other_transport_fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")  # Light purple for other
            other_transport_font = Font(color="8E24AA", bold=True)
            
            border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )
            
            # Function to categorize field types for color coding
            def get_field_category(field_name: str) -> str:
                """Categorize field for color coding"""
                field_lower = field_name.lower()
                
                # Address-related fields
                if any(keyword in field_lower for keyword in [
                    'address', 'street', 'home', 'residence', 'location', 'road', 'avenue', 'lane', 'drive'
                ]):
                    return 'address'
                
                # Student name-related fields
                if any(keyword in field_lower for keyword in [
                    'student', 'name', 'first', 'last', 'full', 'pupil', 'learner'
                ]):
                    return 'student_name'
                
                # Daycare-related fields
                if any(keyword in field_lower for keyword in [
                    'daycare', 'childcare', 'preschool', 'nursery', 'care', 'provider', 'babysitter'
                ]):
                    return 'daycare'
                
                # Default category for other fields
                return 'general'
            
            # Function to get colors based on category
            def get_category_colors(category: str) -> tuple:
                """Get fill and font colors for a category"""
                if category == 'address':
                    return address_fill, address_font
                elif category == 'student_name':
                    return student_name_fill, student_name_font
                elif category == 'daycare':
                    return daycare_fill, daycare_font
                else:
                    return general_fill, general_font
            
            # Format main data sheet
            if "Traversa_Ready_Data" in workbook.sheetnames:
                sheet = workbook["Traversa_Ready_Data"]
                
                # Format headers
                for cell in sheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border
                
                # Highlight changed cells with category-specific colors
                for change in self.changes_detected:
                    field_name = change['field']
                    
                    # Determine the category and get appropriate colors
                    category = get_field_category(field_name)
                    change_fill, change_font = get_category_colors(category)
                    
                    # Find the column index for this field
                    col_index = None
                    for idx, cell in enumerate(sheet[1], 1):
                        if cell.value == field_name:
                            col_index = idx
                            break
                    
                    if col_index:
                        # Find the row in the filtered data
                        for row_idx in range(2, len(self.traversa_data) + 2):
                            cell = sheet.cell(row=row_idx, column=col_index)
                            if str(cell.value) == str(change['new_value']):
                                cell.fill = change_fill
                                cell.font = change_font
                                
                                # Add comment explaining the change with category info
                                comment_text = f"[{category.upper().replace('_', ' ')}] Updated from: {change['old_value']}\nTo: {change['new_value']}"
                                cell.comment = Comment(comment_text, "Traversa Processor")
                                break
                
                # Apply transportation-specific highlighting
                if 'When do you need transportation (check all that apply)' in [cell.value for cell in sheet[1]]:
                    self.logger.info("🎨 Applying transportation-specific color coding...")
                    
                    # Find transportation column
                    transport_col_idx = None
                    for idx, cell in enumerate(sheet[1], 1):
                        if cell.value == 'When do you need transportation (check all that apply)':
                            transport_col_idx = idx
                            break
                    
                    if transport_col_idx:
                        # Color code each transportation cell based on category
                        for row_idx in range(2, len(self.traversa_data) + 2):
                            transport_cell = sheet.cell(row=row_idx, column=transport_col_idx)
                            transport_text = str(transport_cell.value) if transport_cell.value else ""
                            
                            # Analyze transportation needs for this cell
                            transport_category = self._analyze_transportation_needs(transport_text)
                            
                            # Apply appropriate color based on transportation category
                            if transport_category == "AM Only":
                                transport_cell.fill = am_only_fill
                                transport_cell.font = am_only_font
                                comment_text = "🌅 AM Transportation Only - Home to school"
                            elif transport_category == "PM Only":
                                transport_cell.fill = pm_only_fill
                                transport_cell.font = pm_only_font
                                comment_text = "🌆 PM Transportation Only - School to home"
                            elif transport_category == "Both AM & PM":
                                transport_cell.fill = both_transport_fill
                                transport_cell.font = both_transport_font
                                comment_text = "🔄 Both AM & PM Transportation - Full service"
                            elif transport_category == "No Transportation":
                                transport_cell.fill = no_transport_fill
                                transport_cell.font = no_transport_font
                                comment_text = "🚫 No Transportation - Service declined"
                            else:  # Other/Unclear
                                transport_cell.fill = other_transport_fill
                                transport_cell.font = other_transport_font
                                comment_text = "❓ Transportation needs unclear - Review required"
                            
                            # Add comment explaining transportation category
                            transport_cell.comment = Comment(comment_text, "Transportation Analyzer")
                
                # Auto-adjust column widths
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Format other sheets
            for sheet_name in ["Changes_Summary", "Processing_Summary", "Color_Legend"]:
                if sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    
                    # Format headers
                    for cell in sheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = border
                    
                    # Special formatting for Color Legend sheet
                    if sheet_name == "Color_Legend":
                        # Apply actual colors to the Color Legend sheet
                        for row_idx in range(2, 12):  # Updated range for transportation categories
                            if row_idx > sheet.max_row:
                                break
                                
                            category_cell = sheet.cell(row=row_idx, column=1)  # Category column
                            color_cell = sheet.cell(row=row_idx, column=2)     # Color column
                            
                            category_value = category_cell.value
                            if category_value == "Address":
                                color_cell.fill = address_fill
                                color_cell.font = address_font
                            elif category_value == "Student Name":
                                color_cell.fill = student_name_fill
                                color_cell.font = student_name_font
                            elif category_value == "Daycare":
                                color_cell.fill = daycare_fill
                                color_cell.font = daycare_font
                            elif category_value == "General":
                                color_cell.fill = general_fill
                                color_cell.font = general_font
                            elif category_value == "AM Only":
                                color_cell.fill = am_only_fill
                                color_cell.font = am_only_font
                            elif category_value == "PM Only":
                                color_cell.fill = pm_only_fill
                                color_cell.font = pm_only_font
                            elif category_value == "Both AM & PM":
                                color_cell.fill = both_transport_fill
                                color_cell.font = both_transport_font
                            elif category_value == "No Transportation":
                                color_cell.fill = no_transport_fill
                                color_cell.font = no_transport_font
                            elif category_value == "Other/Unclear":
                                color_cell.fill = other_transport_fill
                                color_cell.font = other_transport_font
                    
                    # Auto-adjust column widths
                    for column in sheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        
                        for cell in column:
                            try:
                                if cell.value and len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        adjusted_width = min(max_length + 2, 50)
                        sheet.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(file_path)
            self.logger.info("✅ Traversa Excel formatting applied successfully")
            
        except Exception as e:
            self.logger.error(f"❌ Error applying Traversa formatting: {e}")
    
    def _get_changes_summary(self) -> Dict[str, int]:
        """Get summary of changes by field type and category"""
        field_summary = {}
        category_summary = {}
        
        for change in self.changes_detected:
            # Count by field name
            field = change['field']
            field_summary[field] = field_summary.get(field, 0) + 1
            
            # Count by category
            category = change.get('category', 'general')
            category_summary[category] = category_summary.get(category, 0) + 1
        
        return {
            'by_field': field_summary,
            'by_category': category_summary,
            'total_changes': len(self.changes_detected)
        }


def main():
    """Main function for Traversa data processing"""
    print("🚌 Traversa Data Processor")
    print("=" * 50)
    
    # Initialize processor
    processor = TraversaDataProcessor()
    
    # Get file paths
    ai_file = input("📁 Enter path to AI extracted data file: ").strip()
    if not ai_file:
        ai_file = "/Users/chaytonobando/Library/Mobile Documents/com~apple~CloudDocs/Python/AI_PDF_Data_Extractor/ROCORI_380_Files_Extraction.xlsx"
    
    traversa_template = input("📁 Enter path to Traversa template file: ").strip()
    
    if not traversa_template:
        print("❌ Traversa template file is required")
        return
    
    output_file = input("📁 Enter output file path (or press Enter for auto): ").strip()
    if not output_file:
        output_file = f"traversa_ready_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    try:
        # Optional: Set custom field mappings
        print("\n🔧 Field Mapping Options:")
        print("1. Auto-map fields (recommended)")
        print("2. Manual field mapping")
        
        choice = input("Choose option (1 or 2): ").strip()
        
        if choice == "2":
            print("📋 Enter field mappings (AI_field:Traversa_field, comma separated):")
            print("Example: Student Name:Student_Name,Grade:Grade_Level")
            mapping_input = input().strip()
            
            if mapping_input:
                mappings = {}
                for pair in mapping_input.split(','):
                    if ':' in pair:
                        ai_field, traversa_field = pair.split(':', 1)
                        mappings[ai_field.strip()] = traversa_field.strip()
                
                processor.set_field_mappings(mappings)
        
        # Process data
        results = processor.process_for_traversa(
            ai_extractor_file=ai_file,
            traversa_template_file=traversa_template,
            output_file=output_file,
            fuzzy_threshold=80,
            auto_map_fields=(choice != "2")
        )
        
        print(f"\n✅ Traversa processing complete!")
        print(f"📊 Summary:")
        print(f"   🎓 Original students: {results['total_original_students']}")
        print(f"   ✅ Students kept: {results['matched_students']}")
        print(f"   ❌ Students removed: {results['removed_students']}")
        print(f"   🔄 Fields updated: {results['updated_fields']}")
        print(f"   📁 Output file: {results['output_file']}")
        
        if results['changes_summary']['by_category']:
            print(f"\n🎨 Changes by category:")
            category_names = {'address': '🏠 Address', 'student_name': '👤 Student Name', 'daycare': '🏫 Daycare', 'general': '📝 General'}
            for category, count in results['changes_summary']['by_category'].items():
                display_name = category_names.get(category, category.title())
                print(f"   {display_name}: {count} updates")
        
        if results['changes_summary']['by_field']:
            print(f"\n📋 Changes by field:")
            for field, count in results['changes_summary']['by_field'].items():
                print(f"   🔄 {field}: {count} updates")
        
        print(f"\n🚌 File is ready for Traversa upload!")
        
    except Exception as e:
        print(f"❌ Error: {e}")


if __name__ == "__main__":
    main()
